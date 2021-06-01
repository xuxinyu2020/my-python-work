"""
版本：v1.0
版本说明：调用/imcrs/apm/application/appperflist/{app_id}接口，在旧版本中没有instanceName字段，无法使用。
        测试apm版本（UCenter APM 7.3 (E0707L02)）。在v2.0中会使用接口/imcrs/apm/application/app/{id}，方式更简单。
脚本作者：xuxinyu
脚本最后更新时间：2021/5/29 0:55
脚本说明：
    本脚本用于调取imc apm接口，对监控的服务器磁盘利用率信息进行统计，输出报表。
    适用pyhon3.x版本，执行前需要安装预对应模块。

    脚本执行方法：命令行界面执行python imc_osdisk_statics.py
    脚本执行成功后，会在脚本所在目录下生成excel表格。

    目前可统计的操作系统类型有：windows，linux和AIX

注意事项：频繁调用apm接口会影响性能。考虑现场的应用数量较多，请不要频繁执行此脚本，避免对服务器造成性能压力。
"""
import json
from datetime import datetime

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Side, colors, Font
from openpyxl.utils import get_column_letter
from requests.auth import HTTPDigestAuth
import pandas as pd

"""指标说明：
    winsvr.disk                                                   = 磁盘利用率
    winsvr.disk.DiskName                                          = 磁盘
    winsvr.disk.DiskSize                                          = 容量
    winsvr.disk.DiskFreeSpace                                     = 空闲空间
    winsvr.disk.DiskUsed                                          = 已用空间
    winsvr.disk.DiskUtilization                                   = 磁盘利用率
    winsvr.disk.DiskFreePercentage								  = 磁盘空闲率
    winsvr.disk.DiskAddRatio								  	  = 使用增长率 
    
    linux.disk                                                    = 文件系统
    linux.disk.DiskName                                           = 名称
    linux.disk.DiskSize                                           = 容量
    linux.disk.DiskFreeSpace                                      = 空闲空间
    linux.disk.DiskUsed                                           = 已用空间
    linux.disk.DiskUtilization                                    = 文件系统利用率
    linux.disk.FileSysType                                       = 文件系统类型
    linux.disk.DiskAddRatio										  = 使用增长率

    aix.disk                                                      = 文件系统
    aix.disk.DiskName                                             = 名称
    aix.disk.DiskSize                                             = 容量
    aix.disk.DiskFreeSpace                                        = 空闲空间
    aix.disk.DiskUsed                                             = 已用空间
    aix.disk.DiskUtilization                                      = 文件系统利用率
    aix.disk.DiskAddRatio										  = 使用增长率
    aix.interface                                                 = 网络接口

"""


def StrOfSize(size):
    '''
    auth: wangshengke@kedacom.com ；科达柯大侠
    递归实现，精确为最大单位值 + 小数点后三位
    '''

    def strofsize(integer, remainder, level):
        if integer >= 1024:
            remainder = integer % 1024
            integer //= 1024
            level += 1
            return strofsize(integer, remainder, level)
        else:
            return integer, remainder, level

    units = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
    integer, remainder, level = strofsize(size, 0, 0)
    if level + 1 > len(units):
        level = -1
    return ('{}.{:>03d} {}'.format(integer, remainder, units[level]))


def applist_static():
    """查询imc应用列表，返回一个列表"""
    # 查询参数
    imc_url = 'http://10.153.49.82:8080'  # imc url
    api_url = '/imcrs/apm/application/applist'  # Resful API接口调用
    # 账号和密码
    usernames = 'admin'
    passwords = 'admin'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    # 调用REQUEST GET模块获取信息
    res = requests.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords))
    # print(res.url)
    # 将格式转化为UTF-8

    res.encoding = 'utf-8'
    # 读取为JSON格式方便查询参数
    res_JSON = res.json()
    # 打印json文件
    # with open('applist.json', 'w', encoding='utf-8') as f:
    #     json.dump(res_JSON, f, ensure_ascii=False)  # json.dump()默认用ASCII码解码，注意要加参数，否则不显示中文

    OS_list = []
    for app in res_JSON["app"]:
        if app["type"] in ['winsvr', 'linux', 'aix']:
            OS_list.append({
                "appId": app["appId"],
                "name": app["name"],
                "type": app["type"],
                "ip": app["ip"],
            })
    return OS_list


def disk_static(app_id, fieldname, cellname='disk'):
    """通过API查询单个操作系统磁盘利用率数据"""
    """参数说明：
        app_id imc应用对应的appid
        cellname 指标组名称，默认disk
        fieldname:指标名称"""
    # 已知操作系统有windows，linux和AIX
    # 查询参数
    imc_url = 'http://10.153.49.82:8080'  # uc ip
    api_url = f'/imcrs/apm/application/appperflist/{app_id}'  # Resful API接口调用
    # 账号和密码
    usernames = 'admin'
    passwords = 'admin'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    # paradata = {
    #     "cellname":"disk",
    #     "fieldname":"DiskUtilization",
    # }
    paradata = {
        "cellname": cellname,
        "fieldname": fieldname,
    }

    # 调用REQUEST GET模块获取信息
    res = requests.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords), params=paradata)
    # 将格式转化为UTF-8
    res.encoding = 'utf-8'
    # 读取为JSON格式方便查询参数
    res_JSON = res.json()
    # 打印json文件
    # with open(f'{fieldname}.json','w',encoding='utf-8') as f:
    #     json.dump(res_JSON,f,ensure_ascii=False)#json.dump()默认用ASCII码解码，注意要加参数，否则不显示中文
    return res_JSON


def all_disk_static(filedname):
    """查询所有os的磁盘利用率"""
    OS_list = applist_static()  # 查找os的列表
    # 用于存放返回数据
    result = []
    value_list = []

    for OS in OS_list:
        # print(OS_list)
        res_util = disk_static(OS["appId"], fieldname=filedname)
        value = 0  # 用于存放空闲空间或已使用空间
        # res_util = disk_static(33,fieldname="DiskUtilization")
        instance_list = res_util["fieldList"]["field"]["instanceList"]
        unit = res_util["fieldList"]["field"]['unit']
        # print("instance_list打印输出如下：")
        # print(instance_list)
        # print("instance_list类型输出如下：")
        # print(type(instance_list))
        # print("instance_list['instance']打印输出如下：")
        # print(instance_list["instance"])
        # print("instance_list['instance']类型打印输出如下：")
        # print(type(instance_list["instance"]))
        DiskUtilization = []
        if instance_list is None:
            DiskUtilization.append("无数据")
        elif type(instance_list['instance']) is dict:
            #     如果操作系统只有一个盘，instance_list['instance']返回的是字典
            # 对于空闲磁盘和已用磁盘做单位换算
            if unit == '%':
                value = None
                DiskUtilization.append(instance_list['instance']['instanceName'] + ' ' +
                                       instance_list['instance']['perfDataList']['perfData'][-1]['value'] + unit)
            elif unit == 'KB':
                value = eval(instance_list['instance']['perfDataList']['perfData'][-1]['value'])
                perf_value = StrOfSize(int(eval(instance_list['instance']['perfDataList']['perfData'][-1]['value'])))
                DiskUtilization.append(instance_list['instance']['instanceName'] + ' ' + perf_value)
        elif type(instance_list['instance']) is list:
            #     如果操作系统是多个盘，返回的是包含多实例的list
            # print(instance_list['instance'])
            for instance in instance_list['instance']:
                # print("-----------------------------")
                # print(instance)
                # print(f"instance_list的类型是：{type(instance_list)}")
                # print(f"instance的类型是：{type(instance)}")
                if unit == '%':
                    value = None
                    DiskUtilization.append(
                        instance['instanceName'] + ' ' + instance['perfDataList']['perfData'][-1]['value'] + unit)
                elif unit == 'KB':
                    value += eval(instance['perfDataList']['perfData'][-1]['value'])
                    perf_value = StrOfSize(int(eval(instance['perfDataList']['perfData'][-1]['value'])))
                    DiskUtilization.append(instance['instanceName'] + ' ' + perf_value)

        result.append(DiskUtilization)
        value_list.append(value)

    return result, value_list


def disk_to_xls():
    """定义函数，将数据写入excel表格"""
    # 调用接口查询所有参数，获得列表
    print("开始调用imc接口")
    OS_list = applist_static()  # 查询操作系统列表
    print("获取服务器列表成功，开始获取磁盘使用数据...")
    print("若应用较多，此过程执行较缓慢，请耐心等待...")
    DiskUtil_list, DiskUtil_value = all_disk_static('DiskUtilization')  # 查询磁盘使用率
    DiskUsed_list, DiskUsed_value = all_disk_static('DiskUsed')  # 查询磁盘已用空间
    DiskFree_list, DiskFree_value = all_disk_static('DiskFreeSpace')  # 查询磁盘空闲空间
    print("获取数据成功！")

    print('开始将数据写入表格...')

    OS_num = len(OS_list)
    disk_data = []
    for i in range(OS_num):
        disk_used_value = DiskUsed_value[i]
        disk_free_value = DiskFree_value[i]
        if disk_used_value == 0 and disk_free_value == 0:
            alldisk_util = '无数据'
        else:
            alldisk_util = str('%.2f%%' % (disk_used_value / (disk_free_value + disk_used_value) * 100))

        disk_data.append((OS_list[i]['name'], "\n".join(i for i in DiskUtil_list[i]),
                          "\n".join(i for i in DiskUsed_list[i]), "\n".join(i for i in DiskFree_list[i]), alldisk_util))
    # 将列表写入excel
    data = pd.DataFrame(disk_data, columns=['主机名', '磁盘/文件利用率', '磁盘已用空间', '磁盘空闲空间', '设备整体磁盘/文件使用百分比'])

    filename = "深圳航空iMC服务器磁盘利用率信息统计表"+datetime.now().strftime("%Y%m%d%H%M%S")+".xlsx"


    data.to_excel(filename, startrow=1, index=False)
    print(f"{filename}表格写入成功！")

    dealwith_xls(filename)

    print("Success!请至脚本所在目录下取用表格。")


def dealwith_xls(filename):
    """处理excel表格格式"""
    # 设置单元格格式参数
    align_set = Alignment(horizontal='left',vertical='center', wrapText=True)
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK),
                        )
    # 设置标题格式参数
    title_align_set = Alignment(horizontal='center', vertical='center')
    title_font_set = Font(name=u"宋体", bold='bold', size=14)

    #     打开表格
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook["Sheet1"]

    # 设置标题内容及格式
    worksheet['A1'].value = '深圳航空iMC服务器磁盘利用率信息统计表' + datetime.now().strftime("%Y%m%d")
    worksheet.merge_cells(range_string='A1:E1')
    worksheet['A1'].font = title_font_set
    worksheet['A1'].alignment = title_align_set
    worksheet.row_dimensions[1].height = 60

    #     统计表格最大行和列
    nrows = worksheet.max_row
    ncols = worksheet.max_column

    # 设置单元格自动换行和边框
    for i in range(1,nrows):
        for j in range(ncols):
            worksheet.cell(row=i + 1, column=j + 1).alignment = align_set
            worksheet.cell(row=i + 1, column=j + 1).border = border_set

    # 设置所有行宽均为25
    for i in range(1, ncols + 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25
    #     保存修改
    workbook.save(filename)
    print(f"{filename}格式调整成功")


if __name__ == '__main__':
    # disk_static(33,fieldname="DiskUsed")
    # disk_static(33,fieldname="DiskFreeSpace")
    # applist_static()
    # all_disk_static('DiskUtilization') #查询磁盘使用率
    # all_disk_static('DiskFreeSpace') #查询磁盘空闲空间
    # all_disk_static('DiskUsed') #查询磁盘已用空间
    disk_to_xls()
