import json
from datetime import datetime
import requests as rq
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, colors
from openpyxl.utils import get_column_letter
import openpyxl
from requests.auth import HTTPDigestAuth


def requestConnect(api_url):
    """定义函数通过API接口获取json格式文件。
    参数说明：api_url:api接口
    """
    # 查询参数
    # imc地址
    imc_url = 'http://10.14.136.80'
    # 账号和密码
    usernames = 'admin'
    passwords = 'shenzhen@1qaz2wsx'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    # 调用REQUEST GET模块获取信息
    res = rq.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords))
    # 将格式转化为UTF-8
    res.encoding = 'utf-8'
    # 读取为JSON格式方便查询参数
    res_JSON = res.json()

    return res_JSON


def json_to_excel():
    # 网络设备数量统计
    dev_JSON = requestConnect('/imcrs/plat/res/device?size=1000')
    dev_num = len(dev_JSON['device'])
    app_JSON = requestConnect('/imcrs/apm/application/applist')

    # with open('app.json','w') as f:
    #     json.dump(app_JSON,f)
    # ----------------------------------------------------------------
    # 应用数量总计
    app_num = len(app_JSON['app'])
    # print(f'应用数量是{app_num}')
    # 用于存储应用类型的数量
    app_dict = {
        'storage': 0,
        'unix': 0,
        'vm': 0,
        'windows': 0,
        'appserver': 0,
        'Custom': 0,
        'httpservice': 0,
        'service': 0,
        'linux': 0,
        'db': 0
    }
    for i in app_JSON['app']:
        app_dict[i["typeCategory"]] += 1

    # 各个应用数量统计
    OS_num = app_dict['windows']+app_dict['unix']+app_dict['linux']
    stor_num=app_dict['storage']
    db_num=app_dict['db']
    other_num=app_num-OS_num-stor_num-db_num

    # 用于存储查询数量
    result = [('网络设备', dev_num), ('OS操作系统', OS_num), ('数据库', db_num), ('存储设备', stor_num), ('其他', other_num)]
    data = pd.DataFrame(result, columns=['类型', '数量'])  # 用于存储每一行的Json数据

    filename = "深航imc监控统计表"+datetime.now().strftime("%Y%m%d") + ".xlsx"
    # 在excel表格的第1列写入, 不写入index
    data.to_excel(filename, startrow=1, index=False)
    print(f"{filename}表格输出成功")

    deal_xls(filename)


def deal_xls(filename):
    """excel表格格式处理"""
    # 日志输出
    print("开始调整表格样式...")

    # 定义单元格格式
    font_set = Font(name=u"宋体",bold='bold',size=14)
    align_set = Alignment(horizontal='center',vertical='center')
    border_set = Border(left=Side(style='thin',color=colors.BLACK),
                        right=Side(style='thin',color=colors.BLACK),
                        top=Side(style='thin',color=colors.BLACK),
                        bottom=Side(style='thin',color=colors.BLACK),
                        )
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name("Sheet1")

    # 设置标题：内容，合并单元格，设置样式
    worksheet['A1'].value = '深圳航空iMC监控信息统计表'
    worksheet.merge_cells(range_string='A1:B1')
    worksheet['A1'].alignment = align_set
    worksheet['A1'].font = font_set
    worksheet.row_dimensions[1].height=40

    #单元格样式设置
    nrows = worksheet.max_row
    ncols = worksheet.max_column
    # 设置行宽
    for i in range(1,worksheet.max_column+1):
            worksheet.column_dimensions[get_column_letter(i)].width= 20
    # 居中对齐,给单元格加边框
    for i in range(nrows):
        for j in range(ncols):
            worksheet.cell(row=i+1,column=j+1).alignment = align_set
            worksheet.cell(row=i+1,column=j+1).border = border_set

    # 保存表格
    workbook.save(filename)
    print(f"{filename}样式调整成功！")


if __name__ == '__main__':
    json_to_excel()
