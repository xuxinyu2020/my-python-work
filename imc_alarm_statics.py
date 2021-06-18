import json
import re,time,openpyxl,sys
from datetime import datetime
import requests as rq
import pandas as pd
from requests.auth import HTTPDigestAuth
from openpyxl.styles import Font, Alignment, Border, Side,colors
from openpyxl.utils import get_column_letter

def deal_xls(filename,t1,t2):
    """excel表格格式处理"""
    # 日志输出
    print("开始调整表格样式...")

    # 定义单元格格式
    font_set = Font(name=u"宋体",bold='bold',size=14)
    align_set = Alignment(horizontal='center',vertical='center')
    border_set = Border(left=Side(style='thin',color=colors.BLACK),
                        right=Side(style='thin',color=colors.BLACK),
                        top=Side(style='thin',color=colors.BLACK),
                        bottom=Side(style='thin',color=colors.BLACK),
                        )
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name("Sheet1")

    # 设置标题：内容，合并单元格，设置样式
    worksheet['A1'].value = '深航航空iMC业务系统频繁告警信息统计表'+f"({t1}-{t2})"
    worksheet.merge_cells(range_string='A1:F1')
    worksheet['A1'].alignment = align_set
    worksheet['A1'].font = font_set
    worksheet.row_dimensions[1].height=90

    # 重复单元格合并
    nrows = worksheet.max_row
    ncols = worksheet.max_column
    # 定义起始行m,定义结束列n
    m = 3
    n = 4
    while n <= nrows:
        if worksheet.cell(column=1, row=n).value == worksheet.cell(column=1, row=m).value:
            n += 1
        else:
            worksheet.merge_cells(range_string=f'A{m}:A{n - 1}')
            worksheet.merge_cells(range_string=f'B{m}:B{n - 1}')
            worksheet.merge_cells(range_string=f'C{m}:C{n - 1}')
            # worksheet.merge_cells(range_string=f'D{m}:D{n - 1}')
            m = n
            n += 1
    worksheet.merge_cells(range_string=f'A{m}:A{n - 1}')
    worksheet.merge_cells(range_string=f'B{m}:B{n - 1}')
    worksheet.merge_cells(range_string=f'C{m}:C{n - 1}')
    # worksheet.merge_cells(range_string=f'D{m}:D{n - 1}')

    #单元格样式设置
    # 设置行宽
    for i in range(1,worksheet.max_column+1):
        if i == 5:
            worksheet.column_dimensions[get_column_letter(i)].width= 80
        else:
            worksheet.column_dimensions[get_column_letter(i)].width= 20
    # 居中对齐,给单元格加边框
    for i in range(nrows):
        for j in range(ncols):
            worksheet.cell(row=i+1,column=j+1).alignment = align_set
            worksheet.cell(row=i+1,column=j+1).border = border_set

    # 保存表格
    workbook.save(filename)
    print(f"{filename}样式调整成功！")

def requestConnect(t1,t2):
    """定义函数，用于连接imc获取告警数据，处理后输出excel表格"""
    """参数含义：t1:告警初始时间；t2:结束时间"""
    #查询参数
    imc_url = 'http://10.153.49.82:8080'    # imc访问地址
    api_url = '/imcrs/fault/alarm'     # Resful API接口调用
    # 账号和密码
    usernames = 'admin'
    # passwords = 'shenzhen@1qaz2wsx'
    passwords = 'admin'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    # 告警查询时间
    startAlarmTime= int(time.mktime(time.strptime(t1,"%Y/%m/%d_%H:%M:%S")))
    endAlarmTime= int(time.mktime(time.strptime(t2,"%Y/%m/%d_%H:%M:%S")))
    # paradata = {
    #     "operatorName":"admin",
    #     "alarmLevel":1,
    #     "startAlarmTime":startAlarmTime,
    #     "endAlarmTime":endAlarmTime,
    #     "size":1000000000
    # } 由于查询告警等级有1和2两个，因此paradata采用list，不采用dict，因为dict的key唯一
    paradata=[
        ("operatorName","admin"),
        ("alarmLevel",1),
        ("alarmLevel",2),
        ("startAlarmTime",startAlarmTime),
        ("endAlarmTime",endAlarmTime),
        ("size",1000000000)
    ]
    # 调用REQUEST GET模块获取信息
    print("正在连接imc...")
    res = rq.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords),params=paradata)
    print("获取数据成功，正在处理中...")
    # 将格式转化为UTF-8
    res.encoding = 'utf-8'
    # 读取为JSON格式方便查询参数
    res_JSON = res.json()
    
    # 打印json文件
    # with open('alarm.json','w',encoding='utf-8') as f:
    #     json.dump(res_JSON,f,ensure_ascii=False)#json.dump()默认用ASCII码解码，注意要加参数，否则不显示中文
    # print(res.url)

    # -------------------------------------------------------------
    """测试代码"""
    # """
    # with open(r'test_scripts/shenzhen_air_alarm.json',encoding='utf-8') as json_f:
    #     res_JSON = json.load(json_f)
    #     print(res_JSON)
    #
    # """
    # -------------------------------------------------------------

    # 设备告警分类id
    alarm_catagory_dict={
        "设备告警":[2,3,4,5,7],
        "网管告警":[1901],
        "其他告警":[1,15],
        "应用告警":[16,1191,1261,2701,1093],
        "性能告警":[8,9],
        "服务器告警":[1271],
        "安全告警":[10,11,12,13,1101],
        "配置告警":[14],
        "虚拟化平台":[1291],
        "CIM component":[1120],
        "CMDB配置项废止提前报警":[1801]
    }

    # 生成表格的文件名
    alarm_xls = "深圳航空iMC业务系统频繁告警信息统计表"+datetime.now().strftime("%Y%m%d%H%M%S")+".xlsx"
    # json数据中应用告警的设备名称替换。应用告警名称默认是127.0.0.1,替换为实际应用名称。
    data = pd.DataFrame()  # 用于存储每一行的Json数据
    for static in res_JSON["alarm"]:
        if int(static["alarmCategory"]) in alarm_catagory_dict["应用告警"]:
            # 应用告警告警设备名称替换为实际应用
            appName = re.findall(r'APM Monitor Name=(.*?);',static["paras"])[0]
            appIP = re.findall(r'Device IP=(.*?);',static["paras"])[0]
            static["deviceName"] = appName
            static["deviceIp"] = appIP
        data = data.append(static,ignore_index=True)
    # data.to_excel("imc阶段告警全部列表.xls",startrow=0,index=False)

    # 统计告警ip出现次数
    dict = {}
    for alarm in res_JSON["alarm"]:
        if alarm["deviceIp"] in dict:
            dict[alarm["deviceIp"]]+=1
        else:
            dict[alarm["deviceIp"]]=1
    result = []
    # 告警次数超过3次的写入excel表格
    for key,value in dict.items():
        if dict[key] >= 3:
            for alarm in res_JSON["alarm"]:
                if alarm["deviceIp"] == key:
                    result.append([alarm["deviceIp"],alarm["deviceName"],value,alarm["alarmLevelDesc"],alarm["alarmDesc"],alarm["faultTimeDesc"]])
    result_data = pd.DataFrame(result,columns=['告警IP','系统名称','告警次数','告警类别','告警内容','告警时间'])
    result_data.to_excel(alarm_xls,startrow=1,index=False)

    print("告警表格输出成功！")


    # excel表格格式处理
    deal_xls(alarm_xls,t1,t2)

    return 1


if __name__ == '__main__':
    # requestConnect("2021/5/22_00:00:00","2021/5/22_05:00:00")
    requestConnect(sys.argv[1],sys.argv[2])