"""
脚本作者：xuxinyu
脚本最后更新时间：2021/11/30 15:36
脚本说明：
    本脚本用于统计imc监控的网络设备和应用分类统计，适用python3.x版本，并执行前需要预安装对应模块。
    为了减少接口调用对服务器的负担，目前仅针对深航目前监控的应用种类，如果后续新增应用种类，需要重新修改脚本。

    脚本执行方法：命令行界面执行python imc_monitor_statics_v2.1.py
    脚本执行成功后，会在脚本所在目录下生成excel表格。

    目前统计的分类包括：
        网络设备：包含PC设备
        OS操作系统：windows server/AIX/SOLARIS/HP-UX/Linux
        数据库：SQL SERVER/Mysql/Oracle/Oracle ASM
        存储：DELL EqualLogic/Hitachi VSP/NetApp/EMC ISILON/HP 3PAR/IBM DS系列
        中间件：WebSphere MQ
        服务：Ping Command/URL/TCP Port
        其他：其他应用

"""

import json
from datetime import datetime
import requests as rq
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, colors
from openpyxl.utils import get_column_letter
import openpyxl
from requests.auth import HTTPDigestAuth


def requestConnect(api_url):
    """定义函数通过API接口获取接口响应。
    参数说明：api_url:api接口
    """
    # 查询参数
    # imc地址
    # imc_url = 'http://10.14.136.80'
    imc_url = 'http://10.153.49.82:8080'    # imc访问地址
    # 账号和密码
    usernames = 'admin'
    # passwords = 'shenzhen@1qaz2wsx'
    passwords = 'admin'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    # 设置超时时间
    timeout = 10
    # 调用REQUEST GET模块获取信息
    try:
        res = rq.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords),timeout=timeout)
        if res.status_code != 200:
            print("imc连接错误，请检查连接信息、网络或者imc情况！")
            return -1 #-1表示错误
        # 将格式转化为UTF-8
        res.encoding = 'utf-8'
        return res
    except rq.exceptions.ConnectTimeout:
        print("imc连接超时，请检查网络或者imc连接信息是否有误！")
        return -1




def json_to_excel():
    # 网络设备数量统计
    print("开始连接imc")
    if requestConnect('/imcrs/plat/res/device') == -1:
        return
    print("开始统计网络设备数量...")
    dev_num = int(
        requestConnect('/imcrs/plat/res/device?resPrivilegeFilter=false&desc=false&total=true&exact=false').headers[
            'Total'])
    print("网络设备数量统计成功...")
    # 应用数量总计
    print("开始统计应用数量...")
    app_num = int(requestConnect('/imcrs/apm/application/applist?total=True').headers['Total'])
    # print(f'应用数量是{app_num}')
    # 所有应用种类
    # app_dict={'winsvr': 0, 'aix': 0, 'freebsd': 0, 'openbsd': 0, 'hpux': 0, 'sco_sv': 0, 'solaris': 0, 'macos': 0, 'linux': 0,
    #  'kylin': 0, 'mssql': 0, 'mysql': 0, 'oracle': 0, 'dotnet': 0, 'jboss': 0, 'tomcat': 0, 'lync': 0, 'lync13': 0,
    #  'apache': 0, 'iis': 0, 'ex03': 0, 'ex07': 0, 'ex10': 0, 'ex13': 0, 'osp': 0, 'osp10': 0, 'osp13': 0, 'rest': 0,
    #  'soap': 0, 'urlseq': 0, 'ad': 0, 'file': 0, 'dir': 0, 'snmp': 0, 'ftp': 0, 'sftp': 0, 'vplex': 0, 'psql': 0,
    #  'ibmv5k': 0, 'hadoop': 0, 'dataguard': 0, 'hyperv': 0, 'emcvmx': 0, 'ceph': 0, 'pop3': 0, 'winperf': 0,
    #  'db2v11': 0, 'vm2013': 0, 'rabbit': 0, 'sybase': 0, 'wildfly': 0, 'mysql8': 0, 'ogg': 0, 'ibmsvc': 0, 'vcenter': 0,
    #  'ipmi': 0, 'resin': 0, 'vsp': 0, 'gf': 0, 'memch': 0, 'nginx': 0, 'hmc': 0, 'onestor': 0, 'lotushr': 0,
    #  'hw18500v1': 0, 'ping': 0, 'cache2010': 0, 'informix': 0, 'ldap': 0, 'wmq': 0, 'rocky': 0, 'url': 0, 'oras': 0,
    #  'tuxedo': 0, 'mongo': 0, 'db2': 0, 'hws3900': 0, 'ws': 0, 'netapp': 0, 'xugu': 0, 'equallogic': 0, 'jrt': 0,
    #  'dm': 0, 'dbquery': 0, 'comapp': 0, 'lotus': 0, 'amq': 0, 'tcpport': 0, 'p5730': 0, 'as400': 0, 'php': 0, 'kvm': 0,
    #  'cas': 0, 'emc_isilon': 0, 'sap': 0, 'king': 0, 'ams2500': 0, 'hwfc': 0, 'kubemaster': 0, 'msap2k': 0,
    #  'emcclar': 0, 'hw18500': 0, 'citrix': 0, 'smtp': 0, 'dns': 0, 'power': 0, 'pingcmd': 0, 'ibmv7k': 0, 'cmc': 0,
    #  'vmware': 0, 'comsto': 0, 'sc8k': 0, 'ibmds': 0, 'jetty': 0, 'as5600': 0, 'activepro': 0, 'oradataguard': 0,
    #  'oasm': 0, 'hp3par': 0, 'cache': 0, 'suse': 0, 'ibmf900': 0, 'hana': 0, 'hw9000': 0, 'wl': 0, 'tonglq': 0}

    # 用于存储应用类型的数量，只统计了深航监控的部分
    app_dict = {
        'winsvr': 0,
        'aix': 0,
        'mssql': 0,
        'hpux': 0,
        'solaris': 0,
        'linux': 0,
        'mysql': 0,
        'oracle': 0,
        'oasm': 0,
        'url': 0,
        'tcpport': 0,
        'pingcmd': 0,
        'vsp': 0,
        'equallogic': 0,
        'netapp': 0,
        'emc_isilon': 0,
        'ibmds': 0,
        'hp3par': 0,
        'wmq': 0
    }

    # 应用类型数量统计

    for key, value in app_dict.items():
        app_JSON = requestConnect(f'/imcrs/apm/application/applist/?total=True&type={key}')
        app_JSON.encoding = 'utf-8'
        app_dict[key] = int(app_JSON.headers['Total'])

    # 各个应用数量统计
    OS_num = app_dict['winsvr'] + app_dict['aix'] + app_dict['solaris'] + app_dict['hpux'] + app_dict['linux']
    stor_num = app_dict['equallogic'] + app_dict['vsp'] + app_dict['netapp'] + app_dict['emc_isilon'] + app_dict[
        'hp3par'] + app_dict['ibmds']
    db_num = app_dict['mssql'] + app_dict['mysql'] + app_dict['oracle'] + app_dict['oasm']
    service_num = app_dict['pingcmd'] + app_dict['url'] + app_dict['tcpport']
    middle_num = app_dict['wmq']
    other_num = app_num - OS_num - stor_num - db_num - middle_num - service_num
    print(app_dict)
    print(f"应用数量总计：{app_num}")
    print("应用数量统计成功...")
    # 用于存储查询数量
    result = [('网络设备', dev_num), ('OS操作系统', OS_num), ('数据库', db_num), ('存储设备', stor_num), ('中间件',middle_num),('服务监视', service_num),
              ('其他', other_num)]
    data = pd.DataFrame(result, columns=['类型', '数量'])  # 用于存储每一行的Json数据

    filename = "深航imc监控统计表" + datetime.now().strftime("%Y%m%d") + ".xlsx"
    # 在excel表格的第1列写入, 不写入index
    data.to_excel(filename, startrow=1, index=False)
    print(f"{filename}表格输出成功")

    deal_xls(filename)


def deal_xls(filename):
    """excel表格格式处理"""
    # 日志输出
    print("开始调整表格样式...")

    # 定义单元格格式
    font_set = Font(name=u"宋体", bold='bold', size=14)
    align_set = Alignment(horizontal='center', vertical='center')
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK),
                        )
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name("Sheet1")

    # 设置标题：内容，合并单元格，设置样式
    worksheet['A1'].value = '深圳航空iMC监控信息统计表'
    worksheet.merge_cells(range_string='A1:B1')
    worksheet['A1'].alignment = align_set
    worksheet['A1'].font = font_set
    worksheet.row_dimensions[1].height = 40

    # 单元格样式设置
    nrows = worksheet.max_row
    ncols = worksheet.max_column
    # 设置行宽
    for i in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 20
    # 居中对齐,给单元格加边框
    for i in range(nrows):
        for j in range(ncols):
            worksheet.cell(row=i + 1, column=j + 1).alignment = align_set
            worksheet.cell(row=i + 1, column=j + 1).border = border_set

    # 保存表格
    workbook.save(filename)
    print(f"{filename}样式调整成功！")


if __name__ == '__main__':
    json_to_excel()
