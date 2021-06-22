"""
版本：v2.0
版本说明：在v2.0中会使用接口/imcrs/apm/application/app/{id}，方式更简单。解决了在旧版本APM中无法使用的问题
脚本作者：xuxinyu
脚本更新历史：2021/6/2 13:32
            2021/6/18 11:33 根据深航现场部分应用无法采集到文件利用率优化
脚本说明：
    本脚本用于调取imc apm接口，对监控的服务器磁盘利用率信息进行统计，输出报表。
    适用pyhon3.x版本，执行前需要安装预对应模块。

    脚本执行方法：命令行界面执行python imc_osdisk_statics.py
    脚本执行成功后，会在脚本所在目录下生成excel表格。

    目前可统计的操作系统类型有：windows，linux和AIX

注意事项：频繁调用apm接口会影响性能。考虑现场的应用数量较多，请不要频繁执行此脚本，避免对服务器造成性能压力。
"""
import json
from datetime import datetime

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Side, colors, Font
from openpyxl.utils import get_column_letter
from requests.auth import HTTPDigestAuth
import pandas as pd

"""指标说明：
    winsvr.disk                                                   = 磁盘利用率
    winsvr.disk.DiskName                                          = 磁盘
    winsvr.disk.DiskSize                                          = 容量
    winsvr.disk.DiskFreeSpace                                     = 空闲空间
    winsvr.disk.DiskUsed                                          = 已用空间
    winsvr.disk.DiskUtilization                                   = 磁盘利用率
    winsvr.disk.DiskFreePercentage								  = 磁盘空闲率
    winsvr.disk.DiskAddRatio								  	  = 使用增长率 
    
    linux.disk                                                    = 文件系统
    linux.disk.DiskName                                           = 名称
    linux.disk.DiskSize                                           = 容量
    linux.disk.DiskFreeSpace                                      = 空闲空间
    linux.disk.DiskUsed                                           = 已用空间
    linux.disk.DiskUtilization                                    = 文件系统利用率
    linux.disk.FileSysType                                       = 文件系统类型
    linux.disk.DiskAddRatio										  = 使用增长率

    aix.disk                                                      = 文件系统
    aix.disk.DiskName                                             = 名称
    aix.disk.DiskSize                                             = 容量
    aix.disk.DiskFreeSpace                                        = 空闲空间
    aix.disk.DiskUsed                                             = 已用空间
    aix.disk.DiskUtilization                                      = 文件系统利用率
    aix.disk.DiskAddRatio										  = 使用增长率
    aix.interface                                                 = 网络接口

"""


def StrOfSize(float_size):
    '''
    递归实现，精确为最大单位值 + 小数点后三位
    '''
    size=int(float_size)
    def strofsize(integer, remainder, level):
        if integer >= 1024:
            remainder = integer % 1024
            integer //= 1024
            level += 1
            return strofsize(integer, remainder, level)
        else:
            return integer, remainder, level

    units = ['KB', 'MB', 'GB', 'TB', 'PB']
    integer, remainder, level = strofsize(size, 0, 0)
    if level + 1 > len(units):
        level = -1
    return ('{}.{:>03d} {}'.format(integer, remainder, units[level]))


def applist_static():
    """查询imc应用列表，返回一个列表"""
    # 查询参数
    # imc_url = 'http://10.153.49.82:8080'  # imc url
    imc_url = 'http://10.14.136.80'
    api_url = '/imcrs/apm/application/applist'  # Resful API接口调用
    # 账号和密码
    usernames = 'admin'
    # passwords = 'admin'
    passwords = 'shenzhen@1qaz2wsx'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    # 增加查询参数，设置查询数目。默认start是1000，size是1000
    paradata = {
        'start': 0,
        'size': 2550
    }
    # 调用REQUEST GET模块获取信息
    res = requests.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords),params=paradata)
    # 将格式转化为UTF-8
    res.encoding = 'utf-8'
    print(res.url)
    # 读取为JSON格式方便查询参数
    res_JSON = res.json()
    # 用于存放服务器列表
    OS_list = []
    for app in res_JSON["app"]:
        if app["type"] in ['winsvr', 'linux', 'aix']:
            OS_list.append({
                "appId": app["appId"],
                "name": app["name"],
                "type": app["type"],
                "ip": app["ip"],
            })
    return OS_list


def disk_static(app_id, cellname='disk'):
    """通过API查询单个操作系统磁盘利用率数据"""
    """参数说明：
        app_id imc应用对应的appid
        cellname 指标组名称，默认disk
        fieldname:指标名称"""
    # 已知操作系统有windows，linux和AIX
    # 查询参数
    # imc_url = 'http://10.153.49.82:8080'  # uc ip
    imc_url = 'http://10.14.136.80'
    api_url = f'/imcrs/apm/application/app/{app_id}'  # Resful API接口调用
    # 账号和密码
    usernames = 'admin'
    # passwords = 'admin'
    passwords = 'shenzhen@1qaz2wsx'
    # 要求返回的文档类型
    header = {'accept': 'application/json'}
    # 合并URI
    full_url = imc_url + api_url
    paradata = {
        "cellname": cellname,
    }

    # 调用REQUEST GET模块获取信息
    res = requests.get(url=full_url, headers=header, auth=HTTPDigestAuth(usernames, passwords), params=paradata)
    # 将格式转化为UTF-8
    res.encoding = 'utf-8'
    # 读取为JSON格式方便查询参数
    res_JSON = res.json()
    return res_JSON


def deal_disk_static(app_id):
    """处理单个os的磁盘利用率"""

    res_util = disk_static(app_id, cellname='disk')
#---------测试代码-------------
    # with open(f'{app_id}.json','w',encoding='utf-8') as f:
    #     json.dump(res_util,f,ensure_ascii=False)#json.dump()默认用ASCII码解码，注意要加参数，否则不显示中文
    # print("json文件导出成功！")
# ---------测试代码结束-------------
    # 用于存放返回数据，第一个是应用名，后三个空列表分别是磁盘利用率，磁盘已用空间，磁盘空闲空间
    result = [res_util["name"], [], [], [], []]
    # 用于计算设备总磁盘大小和使用量
    all_disk = 0
    all_disk_used = 0

    if "cellList" not in res_util.keys()  or res_util["cellList"] is None:
        # 无数据处理
        for i in range(1, 5):
            result[i].append("无数据")
    else:
        disk_instance = res_util["cellList"]["cell"]["instanceList"]["instance"]
        if type(disk_instance) is dict:
            #     如果操作系统只有一个盘，disk_instance返回的是字典
            dict_temp = {}
            for field in disk_instance['fieldList']['field']:
                dict_temp[field["name"]] = field["value"]
            # {'DiskName': 'C:', 'DiskUtilization': '21.66', 'DiskUsed': '113497088', 'DiskFreeSpace': '410429440',
            #  'DiskAddRatio': '0.00', 'DiskFreePercentage': '78.34', 'DiskSize': '523926528'}
            # print(dict_temp)
            if dict_temp['DiskSize']=="-":
                for i in range(1, 5):
                    result[i].append("无数据")
            else:
                all_disk += float(dict_temp['DiskSize'])
                all_disk_used += float(dict_temp['DiskUsed'])
                all_disk_util = '%.2f%%' % ((all_disk_used / all_disk) * 100)
                result[1].append(dict_temp['DiskName'] + ' ' + dict_temp['DiskUtilization'] + '%')
                result[2].append(dict_temp['DiskName'] + ' ' + StrOfSize(float(dict_temp['DiskUsed'])))
                result[3].append(dict_temp['DiskName'] + ' ' + StrOfSize(float(dict_temp['DiskFreeSpace'])))
                result[4].append(all_disk_util)

        elif type(disk_instance) is list:
            #     如果操作系统是多个盘，返回的是包含多字典的list
            for field_list in disk_instance:
                # 多个磁盘实例
                dict_temp = {}
                for field in field_list['fieldList']['field']:
                    dict_temp[field["name"]] = field["value"]
                # print(dict_temp)
                if dict_temp['DiskSize'] == "-":
                    for i in range(1, 5):
                        result[i].append("无数据")
                else:
                    all_disk += float(dict_temp['DiskSize'])
                    all_disk_used += float(dict_temp['DiskUsed'])
                    result[1].append(dict_temp['DiskName'] + ' ' + dict_temp['DiskUtilization'] + '%')
                    result[2].append(dict_temp['DiskName'] + ' ' + StrOfSize(float(dict_temp['DiskUsed'])))
                    result[3].append(dict_temp['DiskName'] + ' ' + StrOfSize(float(dict_temp['DiskFreeSpace'])))
            # 计算所有盘的磁盘利用率
            if all_disk != 0:
                all_disk_util = '%.2f%%' % ((all_disk_used / all_disk) * 100)
                result[4].append(all_disk_util)

    return result


def disk_to_xls():
    """定义函数，将数据写入excel表格"""
    # 调用接口查询所有参数，获得列表
    print("开始调用imc接口")

    OS_list = applist_static()  # 查询操作系统列表
    disk_result = []  # 要写入excel表格的列表

    print("获取服务器列表成功，开始获取磁盘使用数据...")
    print("若应用较多，此过程执行较缓慢，请耐心等待...")

    for OS in OS_list:
        # 分别获取每台服务器的磁盘列表，构建每一行的数据列表disk_result
        disk_static = deal_disk_static(OS["appId"])
        disk_result.append((disk_static[0], "\n".join(i for i in disk_static[1]), "\n".join(i for i in disk_static[2]),
                            "\n".join(i for i in disk_static[3]), disk_static[4][0]))
    # print(disk_result)
    # 列表写入excel
    data = pd.DataFrame(disk_result, columns=['主机名', '磁盘/文件利用率', '磁盘已用空间', '磁盘空闲空间', '设备整体磁盘/文件使用百分比'])
    filename = "深圳航空iMC服务器磁盘利用率信息统计表" + datetime.now().strftime("%Y%m%d%H%M%S") + ".xlsx"
    data.to_excel(filename, startrow=1, index=False)
    print(f"{filename}表格写入成功！")

    # 调节表格格式
    dealwith_xls(filename)

    print("Success!请至脚本所在目录下取用表格。")


def dealwith_xls(filename):
    """处理excel表格格式"""
    # 设置单元格格式参数
    align_set = Alignment(horizontal='left', vertical='center', wrapText=True)
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK),
                        )
    # 设置标题格式参数
    title_align_set = Alignment(horizontal='center', vertical='center')
    title_font_set = Font(name=u"宋体", bold='bold', size=14)

    #     打开表格
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook["Sheet1"]

    # 设置标题内容及格式
    worksheet['A1'].value = '深圳航空iMC服务器磁盘利用率信息统计表' + datetime.now().strftime("%Y%m%d")
    worksheet.merge_cells(range_string='A1:E1')
    worksheet['A1'].font = title_font_set
    worksheet['A1'].alignment = title_align_set
    worksheet.row_dimensions[1].height = 60

    #     统计表格最大行和列
    nrows = worksheet.max_row
    ncols = worksheet.max_column

    # 设置单元格自动换行和边框
    for i in range(1, nrows):
        for j in range(ncols):
            worksheet.cell(row=i + 1, column=j + 1).alignment = align_set
            worksheet.cell(row=i + 1, column=j + 1).border = border_set

    # 设置所有行宽均为25
    for i in range(1, ncols + 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 25
    #     保存修改
    workbook.save(filename)
    print(f"{filename}格式调整成功")


if __name__ == '__main__':
    disk_to_xls()
